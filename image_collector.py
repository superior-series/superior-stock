import os
from datetime import datetime

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

XLSX_FILE = "image_urls.xlsx"
HEADERS_ROW = ["date", "source", "keyword", "image_url"]


# ---------- 共通ヘルパー ----------

def excel_safe(value):
    """数式インジェクション対策: 数式として解釈されうる先頭文字を無害化する"""
    s = str(value)
    if s and s[0] in ("=", "+", "-", "@"):
        s = "'" + s
    return s


def load_seen(keyword):
    """既存のExcelから『同じキーワード』のURLを読み込み、重複チェック用のsetを返す"""
    seen = set()
    if not os.path.exists(XLSX_FILE):
        return seen
    wb = load_workbook(XLSX_FILE, read_only=True)
    ws = wb.active
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:  # ヘッダー行を飛ばす
            first = False
            continue
        if not row or len(row) < 4:
            continue
        kw, url = row[2], row[3]
        if url and kw in (keyword, excel_safe(keyword)):
            seen.add(url)
    wb.close()
    return seen


def _format_sheet(ws):
    """ヘッダーを太字＋枠固定し、列幅を内容に合わせて整える"""
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None),
                      default=0)
        ws.column_dimensions[letter].width = min(max_len + 2, 80)


def save_urls(source_label, keyword, urls):
    """取得したURLをExcelに追記する（image_url列はクリックできるリンクにする）"""
    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "images"
        ws.append(HEADERS_ROW)

    today = datetime.now().strftime("%Y-%m-%d")
    for url in urls:
        ws.append([today, source_label, excel_safe(keyword), url])
        link_cell = ws.cell(row=ws.max_row, column=4)
        link_cell.hyperlink = url
        link_cell.font = Font(color="0563C1", underline="single")

    _format_sheet(ws)
    wb.save(XLSX_FILE)


# ---------- 取得元ごとの処理 ----------

def scrape_unsplash(keyword, limit, seen, api_key=""):
    """Unsplash API で画像検索し、regular サイズの URL を返す"""
    urls = []
    page = 1
    per_page = min(limit, 30)  # Unsplash API の1リクエスト上限は30

    while len(urls) < limit:
        try:
            resp = requests.get(
                "https://api.unsplash.com/search/photos",
                params={"query": keyword, "per_page": per_page, "page": page},
                headers={"Authorization": f"Client-ID {api_key}"},
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as e:
            print(f"Unsplash APIへの接続に失敗しました: {e}")
            break

        results = data.get("results", [])
        if not results:
            break

        for photo in results:
            url = photo.get("urls", {}).get("regular")
            if not url or url in seen:
                continue
            seen.add(url)
            urls.append(url)
            if len(urls) >= limit:
                break

        print(f"取得済み: {len(urls)}件")

        if len(results) < per_page:  # 最終ページ
            break
        page += 1

    return urls


def scrape_pexels(keyword, limit, seen, api_key=""):
    """Pexels API で画像検索し、large サイズの URL を返す"""
    urls = []
    page = 1
    per_page = min(limit, 80)  # Pexels API の1リクエスト上限は80

    while len(urls) < limit:
        try:
            resp = requests.get(
                "https://api.pexels.com/v1/search",
                params={"query": keyword, "per_page": per_page, "page": page},
                headers={"Authorization": api_key},
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as e:
            print(f"Pexels APIへの接続に失敗しました: {e}")
            break

        photos = data.get("photos", [])
        if not photos:
            break

        for photo in photos:
            url = photo.get("src", {}).get("large")
            if not url or url in seen:
                continue
            seen.add(url)
            urls.append(url)
            if len(urls) >= limit:
                break

        print(f"取得済み: {len(urls)}件")

        if not data.get("next_page"):  # 最終ページ
            break
        page += 1

    return urls


def scrape_pixabay(keyword, limit, seen, api_key=""):
    """Pixabay API で画像検索し、webformatURL を返す"""
    urls = []
    if not api_key:
        return urls
    try:
        params = {
            "key": api_key,
            "q": keyword,
            "image_type": "photo",
            "per_page": max(3, min(200, limit * 3)),  # 3〜200の範囲、重複を考慮して多めに
            "safesearch": "true",
        }
        r = requests.get("https://pixabay.com/api/", params=params, timeout=15)
        if r.status_code != 200:
            return urls
        data = r.json()
        for hit in data.get("hits", []):
            url = hit.get("largeImageURL") or hit.get("webformatURL")
            if not url:
                continue
            if url in seen:
                continue
            urls.append(url)
            if len(urls) >= limit:
                break
    except Exception:
        pass
    return urls


# ---------- メイン ----------

SOURCES = {
    "1": ("Unsplash", "unsplash", scrape_unsplash),
    "2": ("Pexels", "pexels", scrape_pexels),
    "3": ("Pixabay", "pixabay", scrape_pixabay),
}


def main():
    keyword = input("検索キーワードを入力してください: ")

    # 取得件数（1以上の整数になるまで聞き直す）
    while True:
        raw = input("取得件数を入力してください: ")
        try:
            limit = int(raw)
            if limit > 0:
                break
        except ValueError:
            pass
        print("1以上の整数を入力してください。")

    # 取得元の選択
    print("取得元を選んでください:")
    for key, (name, _label, _func) in SOURCES.items():
        print(f"  {key}: {name}")
    valid = "/".join(SOURCES.keys())
    while True:
        choice = input(f"番号を入力 ({valid}): ").strip()
        if choice in SOURCES:
            break
        print(f"{valid} のいずれかを入力してください。")

    name, label, scraper = SOURCES[choice]

    seen = load_seen(keyword)
    if seen:
        print(f"同じキーワードの既存URL {len(seen)}件を重複チェック対象に読み込みました。")

    print(f"{name} から「{keyword}」の画像URLを取得します...")
    urls = scraper(keyword, limit, seen)[:limit]

    save_urls(label, keyword, urls)
    print(f"{len(urls)}件の画像URLを {XLSX_FILE} に保存しました。（取得元: {name}）")


if __name__ == "__main__":
    main()
