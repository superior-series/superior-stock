import os
from io import BytesIO

import pandas as pd
import requests as req
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from image_collector import SOURCES, load_seen, save_urls, XLSX_FILE

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Excel rewrite (used by delete)
# ---------------------------------------------------------------------------

def _rewrite_xlsx(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "images"
    ws.append(["date", "source", "keyword", "image_url"])
    for _, row in df.iterrows():
        ws.append([str(row["date"]), str(row["source"]),
                   str(row["keyword"]), str(row["image_url"])])
        cell = ws.cell(row=ws.max_row, column=4)
        url = str(row["image_url"])
        if url.startswith("http"):
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 80)
    wb.save(XLSX_FILE)


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    sources = [name for _key, (name, _label, _func) in SOURCES.items()]
    return render_template("index.html", sources=sources)


@app.route("/about")
def about():
    return render_template("about.html")


# ---------------------------------------------------------------------------
# API: collect
# ---------------------------------------------------------------------------

@app.route("/api/collect", methods=["POST"])
def collect():
    data = request.get_json(silent=True) or {}
    keyword = (data.get("keyword") or "").strip()
    limit = max(1, min(int(data.get("limit") or 10), 100))
    source_name = (data.get("source") or "").strip()

    if not keyword:
        return jsonify({"ok": False, "message": "キーワードを入力してください"}), 400

    name_to_key = {name: key for key, (name, _label, _func) in SOURCES.items()}
    if source_name not in name_to_key:
        return jsonify({"ok": False, "message": "取得元が不正です"}), 400

    key = name_to_key[source_name]
    name, label, scraper = SOURCES[key]

    _api_key_map = {
        "Unsplash": ("unsplash",
                     "UnsplashのAPIキーが設定されていません。⚙️ API設定からキーを保存してください。"),
        "Pexels":   ("pexels",
                     "PexelsのAPIキーが設定されていません。⚙️ API設定からキーを保存してください。"),
        "Pixabay":  ("pixabay",
                     "PixabayのAPIキーが設定されていません。⚙️ API設定からキーを保存してください。"),
    }
    api_key = ""
    if source_name in _api_key_map:
        short_name, err_msg = _api_key_map[source_name]
        api_key = data.get("api_keys", {}).get(short_name, "")
        if not api_key:
            return jsonify({"ok": False, "message": err_msg}), 400

    seen = load_seen(keyword)
    try:
        if api_key:
            urls = scraper(keyword, limit, seen, api_key=api_key)[:limit]
        else:
            urls = scraper(keyword, limit, seen)[:limit]
        save_urls(label, keyword, urls)
    except Exception as e:
        return jsonify({"ok": False, "message": f"取得中にエラーが発生しました: {e}"}), 500

    if not urls:
        return jsonify({
            "ok": True,
            "count": 0,
            "message": "新しく取得できた画像URLはありませんでした（すべて重複または取得失敗）",
        })

    return jsonify({
        "ok": True,
        "count": len(urls),
        "message": f"{len(urls)} 件を取得して {XLSX_FILE} に保存しました。（取得元: {name}）",
    })


# ---------------------------------------------------------------------------
# API: data (recent 10 rows)
# ---------------------------------------------------------------------------

@app.route("/api/data")
def get_data():
    if not os.path.exists(XLSX_FILE):
        return jsonify({"rows": [], "total": 0})
    try:
        df = pd.read_excel(XLSX_FILE, engine="openpyxl")
    except Exception:
        return jsonify({"rows": [], "total": 0})
    if df.empty:
        return jsonify({"rows": [], "total": 0})

    total = len(df)
    recent = df.tail(10).iloc[::-1]
    rows = []
    for orig_idx, row in recent.iterrows():
        rows.append({
            "idx": int(orig_idx),
            "date": str(row.get("date", "")),
            "source": str(row.get("source", "")),
            "keyword": str(row.get("keyword", "")),
            "image_url": str(row.get("image_url", "")),
        })
    return jsonify({"rows": rows, "total": total})


# ---------------------------------------------------------------------------
# API: delete rows
# ---------------------------------------------------------------------------

@app.route("/api/delete", methods=["POST"])
def delete_rows():
    data = request.get_json(silent=True) or {}
    indices = data.get("indices", [])
    if not indices:
        return jsonify({"ok": False, "message": "削除する行を選択してください"}), 400
    if not os.path.exists(XLSX_FILE):
        return jsonify({"ok": False, "message": "ファイルが見つかりません"}), 404
    try:
        df = pd.read_excel(XLSX_FILE, engine="openpyxl")
        df_updated = df.drop(indices).reset_index(drop=True)
        _rewrite_xlsx(df_updated)
    except Exception as e:
        return jsonify({"ok": False, "message": str(e)}), 500
    return jsonify({"ok": True, "deleted": len(indices)})


# ---------------------------------------------------------------------------
# API: download Excel
# ---------------------------------------------------------------------------

@app.route("/api/download")
def download():
    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "ファイルが見つかりません"}), 404
    return send_file(
        XLSX_FILE,
        as_attachment=True,
        download_name=XLSX_FILE,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# API: download single image (server-side proxy for CORS)
# ---------------------------------------------------------------------------

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg", ".bmp", ".avif"}
_CT_EXT_MAP = {
    "image/jpeg": ".jpg", "image/png": ".png", "image/gif": ".gif",
    "image/webp": ".webp", "image/svg+xml": ".svg", "image/bmp": ".bmp",
    "image/avif": ".avif",
}


@app.route("/api/download-image", methods=["POST"])
def download_image():
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    keyword = (data.get("keyword") or "").strip()
    row_num = int(data.get("row_num") or 1)

    if not url or not url.startswith("http"):
        return jsonify({"error": "URLが不正です"}), 400

    try:
        resp = req.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
    except Exception as e:
        return jsonify({"error": f"画像の取得に失敗しました: {e}"}), 502

    content_type = resp.headers.get("Content-Type", "image/jpeg").split(";")[0].strip()

    # Determine extension: try URL path first, then Content-Type
    url_path = url.split("?")[0]
    url_filename = url_path.rsplit("/", 1)[-1]
    ext = ""
    if "." in url_filename:
        candidate = "." + url_filename.rsplit(".", 1)[-1].lower()
        if candidate in _IMAGE_EXTS:
            ext = candidate
    if not ext:
        ext = _CT_EXT_MAP.get(content_type, ".jpg")

    safe_keyword = keyword.lower().replace(" ", "-") if keyword else "image"
    filename = f"{safe_keyword}_{row_num:03d}{ext}"

    return send_file(
        BytesIO(resp.content),
        as_attachment=True,
        download_name=filename,
        mimetype=content_type or "image/jpeg",
    )


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
